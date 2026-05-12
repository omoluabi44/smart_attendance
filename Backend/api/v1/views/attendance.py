#!/usr/bin/python3
"""Attendance API endpoints — face registration, verification, video/image attendance"""
from models.university import University
from models.attendance import Attendance
from models.attendance_log import AttendanceLog
from models.session import Sessions
from models.session_enrollment import SessionEnrollment
from models.course import Courses
from models.notification import Notification
from models.user import User
from models.lecturer_session import LecturerSession
from models import storage
from api.v1.views import app_views
from flask import abort, jsonify, make_response, request, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from api.v1.utils.roles import role_required
from flasgger.utils import swag_from
from os.path import join, dirname
from datetime import datetime, timedelta
from PIL import Image
import boto3
import os
import io
import uuid
import pandas as pd
from openpyxl.styles import Font

rekognition = boto3.client(
    'rekognition',
    region_name='us-east-1',
    aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
    aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY')
)

s3_client = boto3.client(
    's3',
    region_name='us-east-1',
    aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
    aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY')
)
BUCKET_NAME = 'facepass-file'

ses = storage._DBStorage__session


# ─── Presigned URL Generation ──────────────────────────────────────────────────

@app_views.route('/lecturer/sessions/<session_id>/attendance/presigned-url', methods=['GET'], strict_slashes=False)
@jwt_required()
@role_required('lecturer', 'admin')
def generate_presigned_url(session_id):
    """Generate a presigned URL to upload video/image directly to S3."""
    session_obj = storage.get_id(Sessions, session_id)
    if not session_obj:
        return make_response(jsonify({"error": "Session not found"}), 404)

    file_name = request.args.get('file_name')
    file_type = request.args.get('file_type')

    if not file_name or not file_type:
        return make_response(jsonify({"error": "file_name and file_type are required"}), 400)

    # Validate lecturer is assigned
    lecturer_id = get_jwt_identity()
    is_assigned = ses.query(LecturerSession).filter_by(
        session_id=session_id, lecturer_id=lecturer_id
    ).first()
    if not is_assigned:
        return make_response(jsonify({"error": "You are not assigned to this session"}), 403)

    ext = os.path.splitext(file_name)[1]
    
    # Organize by session id and file type
    if file_type.startswith('video'):
        folder = "attendance_videos"
    else:
        folder = "attendance_images"

    s3_key = f"{folder}/{session_id}/{uuid.uuid4().hex}{ext}"

    try:
        presigned_url = s3_client.generate_presigned_url(
            'put_object',
            Params={
                'Bucket': BUCKET_NAME,
                'Key': s3_key,
                'ContentType': file_type
            },
            ExpiresIn=3600
        )
        return make_response(jsonify({
            "presigned_url": presigned_url,
            "s3_key": s3_key
        }), 200)
    except Exception as e:
        return make_response(jsonify({"error": str(e)}), 500)


# ─── Video Attendance ──────────────────────────────────────────────────────────

@app_views.route('/lecturer/sessions/<session_id>/attendance/video',
                 methods=["POST"], strict_slashes=False)
@jwt_required()
@role_required('lecturer', 'admin')
def upload_attendance_video(session_id):
    """
    Process a video for face-based attendance already uploaded to S3.
    Returns a job_id to poll for results.
    """
    # Validate session exists
    session_obj = storage.get_id(Sessions, session_id)
    if not session_obj:
        return make_response(jsonify({"error": "Session not found"}), 404)

    # Verify lecturer is assigned to this session
    lecturer_id = get_jwt_identity()
    is_assigned = ses.query(LecturerSession).filter_by(
        session_id=session_id, lecturer_id=lecturer_id
    ).first()
    if not is_assigned:
        return make_response(jsonify({"error": "You are not assigned to this session"}), 403)

    data = request.get_json()
    if not data or 's3_key' not in data:
        return make_response(jsonify({"error": "No s3_key provided"}), 400)

    s3_key = data['s3_key']

    try:
        response = rekognition.start_face_search(
            Video={
                'S3Object': {
                    'Bucket': BUCKET_NAME,
                    'Name': s3_key
                }
            },
            CollectionId='Student',
            FaceMatchThreshold=90
        )

        job_id = response['JobId']

        return make_response(jsonify({
            "message": "Video uploaded and face recognition started",
            "job_id": job_id,
            "s3_key": s3_key,
        }), 200)

    except Exception as e:
        return make_response(jsonify({"error": str(e)}), 500)


@app_views.route('/lecturer/sessions/<session_id>/attendance/video-status/<job_id>',
                 methods=['GET'], strict_slashes=False)
@jwt_required()
@role_required('lecturer', 'admin')
def check_video_attendance_status(session_id, job_id):
    """
    Poll the status of a video attendance job.
    When complete, deduplicates faces and records attendance.
    """
    session_obj = storage.get_id(Sessions, session_id)
    if not session_obj:
        return make_response(jsonify({"error": "Session not found"}), 404)

    lecturer_id = get_jwt_identity()

    try:
        response = rekognition.get_face_search(JobId=job_id)
        status = response['JobStatus']

        if status == 'IN_PROGRESS':
            return make_response(jsonify({"status": "processing"}), 200)

        elif status == 'SUCCEEDED':
            # Collect all pages of results
            all_persons = list(response.get('Persons', []))
            next_token = response.get('NextToken')

            while next_token:
                page = rekognition.get_face_search(JobId=job_id, NextToken=next_token)
                all_persons.extend(page.get('Persons', []))
                next_token = page.get('NextToken')

            # ── DEDUPLICATE: keep highest confidence per unique user_id ──
            best_matches = {}
            unknown_count = 0

            for person in all_persons:
                face_matches = person.get('FaceMatches', [])
                if face_matches:
                    best_match = face_matches[0]
                    user_id = best_match.get('Face', {}).get('ExternalImageId')
                    confidence = best_match.get('Similarity', 0)

                    if user_id:
                        if user_id not in best_matches or confidence > best_matches[user_id]:
                            best_matches[user_id] = confidence
                    else:
                        unknown_count += 1
                else:
                    unknown_count += 1

            # Build deduplicated results list
            unique_results = [
                {"user_id": uid, "confidence": conf}
                for uid, conf in best_matches.items()
            ]

            # Record attendance for each unique recognized student
            record_result = update_student_record(unique_results, session_id=session_id, lecturer_id=lecturer_id)

            # Build response with student info
            recognized_students = get_user(unique_results, session_id=session_id)

            return make_response(jsonify({
                "status": "complete",
                "total_frames_analyzed": len(all_persons),
                "unique_faces_recognized": len(unique_results),
                "new_records": record_result["recognized_count"],
                "skipped_already_recorded": record_result["skipped_count"],
                "skipped_students": record_result["skipped_users"],
                "unknown_faces": unknown_count,
                "data": recognized_students,
            }), 200)

        else:
            return make_response(jsonify({
                "status": "failed",
                "message": response.get('StatusMessage', 'Unknown error')
            }), 200)

    except Exception as e:
        return make_response(jsonify({"error": str(e)}), 500)


# ─── Image (Group Photo) Attendance ───────────────────────────────────────────

@app_views.route('/lecturer/sessions/<session_id>/attendance/image',
                 methods=['POST'], strict_slashes=False)
@jwt_required()
@role_required('lecturer', 'admin')
def take_attendance_image(session_id):
    """
    Process a group photo (uploaded to S3) to take attendance via face recognition.
    Detects all faces, identifies each one, and records attendance.
    """
    session_obj = storage.get_id(Sessions, session_id)
    if not session_obj:
        return make_response(jsonify({"error": "Session not found"}), 404)

    lecturer_id = get_jwt_identity()

    # Verify lecturer is assigned
    is_assigned = ses.query(LecturerSession).filter_by(
        session_id=session_id, lecturer_id=lecturer_id
    ).first()
    if not is_assigned:
        return make_response(jsonify({"error": "You are not assigned to this session"}), 403)

    data = request.get_json()
    if not data or 's3_key' not in data:
        return make_response(jsonify({"error": "No s3_key provided"}), 400)

    s3_key = data['s3_key']

    try:
        # Fetch image bytes from S3
        s3_object = s3_client.get_object(Bucket=BUCKET_NAME, Key=s3_key)
        image_bytes = s3_object['Body'].read()

        # Detect all faces in the image
        detection_response = rekognition.detect_faces(
            Image={'Bytes': image_bytes},
            Attributes=['DEFAULT']
        )

        img = Image.open(io.BytesIO(image_bytes))
        width, height = img.size

        all_results = []
        best_matches = {}

        for faceDetail in detection_response['FaceDetails']:
            box = faceDetail['BoundingBox']
            left = width * box['Left']
            top = height * box['Top']
            right = left + (width * box['Width'])
            bottom = top + (height * box['Height'])

            face_crop = img.crop((left, top, right, bottom))

            stream = io.BytesIO()
            face_crop.save(stream, format="JPEG")
            crop_bytes = stream.getvalue()

            try:
                search_response = rekognition.search_users_by_image(
                    CollectionId='Student',
                    Image={'Bytes': crop_bytes},
                    MaxUsers=1,
                    UserMatchThreshold=90
                )
            except rekognition.exceptions.InvalidParameterException:
                continue

            matches = search_response.get('UserMatches', [])
            if matches:
                user_id = matches[0]['User'].get('UserId')
                confidence = matches[0].get('Similarity', 0)
                if user_id and (user_id not in best_matches or confidence > best_matches[user_id]):
                    best_matches[user_id] = confidence

        # Build deduplicated results
        unique_results = [
            {"user_id": uid, "confidence": conf}
            for uid, conf in best_matches.items()
        ]

        # Record attendance
        record_result = update_student_record(unique_results, session_id=session_id, lecturer_id=lecturer_id)

        # Build response
        recognized_students = get_user(unique_results, session_id=session_id)

        return make_response(jsonify({
            "status": "complete",
            "total_faces_detected": len(detection_response['FaceDetails']),
            "unique_faces_recognized": len(unique_results),
            "new_records": record_result["recognized_count"],
            "skipped_already_recorded": record_result["skipped_count"],
            "skipped_students": record_result["skipped_users"],
            "data": recognized_students,
        }), 200)

    except Exception as e:
        return make_response(jsonify({"error": str(e)}), 500)


# ─── Manual Attendance Entry ──────────────────────────────────────────────────

@app_views.route('/attendance', methods=["POST"], strict_slashes=False)
@jwt_required()
@role_required('lecturer', 'admin')
def post_attendance():
    """Create a manual attendance record for a student in a session."""
    data = request.get_json()
    if not data:
        return make_response(jsonify({"error": "Not a JSON"}), 400)

    required_fields = ["session_id", "user_id", "status"]
    for field in required_fields:
        if field not in data:
            return make_response(jsonify({"error": f"Missing required field: {field}"}), 400)

    status = data["status"]
    if status not in ('present', 'absent', 'excused'):
        return make_response(jsonify({"error": "Invalid status. Must be present, absent, or excused"}), 400)

    # Validate session exists
    session_obj = storage.get_id(Sessions, data["session_id"])
    if not session_obj:
        return make_response(jsonify({"error": "Session not found"}), 404)

    # Validate user exists
    user = storage.get_id(User, data["user_id"])
    if not user:
        return make_response(jsonify({"error": "User not found"}), 404)

    lecturer_id = get_jwt_identity()

    # Check for duplicate log today
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    existing_log = ses.query(AttendanceLog).filter(
        AttendanceLog.session_id == data["session_id"],
        AttendanceLog.user_id == data["user_id"],
        AttendanceLog.date >= today_start
    ).first()

    if existing_log:
        existing_log.status = status
        existing_log.recorded_by = lecturer_id
        existing_log.save()
    else:
        log = AttendanceLog(
            session_id=data["session_id"],
            user_id=data["user_id"],
            recorded_by=lecturer_id,
            date=datetime.utcnow(),
            status=status
        )
        log.save()

    # Update summary
    _recalculate_attendance_summary(data["session_id"], data["user_id"])

    return make_response(jsonify({"message": "Attendance recorded"}), 201)


# ─── Face Registration ────────────────────────────────────────────────────────

@app_views.route('/register-face', methods=["POST"], strict_slashes=False)
# @jwt_required()
def register_face():
    """Register a single face image for a user in the Rekognition collection."""
    file = request.files.get('image')
    user_id = request.form.get('user_id')

    if not file:
        return make_response(jsonify({"error": "No image provided"}), 400)
    if not user_id:
        return make_response(jsonify({"error": "No user_id provided"}), 400)

    # Validate user exists in database
    user = storage.get_id(User, user_id.strip())
    if not user:
        return make_response(jsonify({"error": "User not found in database"}), 404)

    try:
        response = rekognition.index_faces(
            CollectionId='Student',
            Image={'Bytes': file.read()},
            ExternalImageId=user_id.strip(),
            MaxFaces=1,
            QualityFilter="AUTO"
        )

        if not response['FaceRecords']:
            return make_response(jsonify({"status": "failed", "message": "No face detected"}), 400)

        face_id = response['FaceRecords'][0]['Face']['FaceId']

        return make_response(jsonify({
            "status": "success",
            "message": f"Face indexed for user: {user_id}",
            "face_id": face_id
        }), 200)

    except Exception as e:
        return make_response(jsonify({"error": str(e)}), 500)


@app_views.route('/register-faces', methods=["POST"], strict_slashes=False)
# @jwt_required()
def register_faces():
    """Register multiple face images for a user (better accuracy)."""
    files = request.files.getlist('images')
    user_id = request.form.get('user_id')

    if not files:
        return make_response(jsonify({"error": "No images provided"}), 400)
    if not user_id:
        return make_response(jsonify({"error": "No user_id provided"}), 400)

    # Validate user exists
    user = storage.get_id(User, user_id.strip())
    if not user:
        return make_response(jsonify({"error": "User not found in database"}), 404)

    face_ids = []
    try:
        # Create Rekognition User (ignore if already exists)
        try:
            rekognition.create_user(CollectionId='Student', UserId=user_id)
        except (rekognition.exceptions.InvalidParameterException,
                rekognition.exceptions.ConflictException):
            pass  # User already exists

        # Index images with ExternalImageId for video compatibility
        for file in files:
            response = rekognition.index_faces(
                CollectionId='Student',
                Image={'Bytes': file.read()},
                ExternalImageId=user_id,
                MaxFaces=1,
                QualityFilter="AUTO"
            )
            if response['FaceRecords']:
                face_ids.append(response['FaceRecords'][0]['Face']['FaceId'])

        # Associate faces with the User API user
        if face_ids:
            rekognition.associate_faces(
                CollectionId='Student',
                UserId=user_id,
                FaceIds=face_ids,
                UserMatchThreshold=70
            )

        return make_response(jsonify({
            "status": "success",
            "message": f"Associated {len(face_ids)} faces to User: {user_id}",
            "face_ids": face_ids
        }), 200)

    except Exception as e:
        return make_response(jsonify({"error": str(e)}), 500)


# ─── Face Verification (Single) ───────────────────────────────────────────────

@app_views.route('/verify-face', methods=["POST"], strict_slashes=False)
@jwt_required()
def verify_face():
    """Verify a single face against the collection."""
    if 'image' not in request.files:
        return make_response(jsonify({"error": "No image uploaded"}), 400)

    file = request.files['image']
    image_bytes = file.read()

    try:
        response = rekognition.search_users_by_image(
            CollectionId='Student',
            Image={'Bytes': image_bytes},
            MaxUsers=1,
            UserMatchThreshold=95
        )

        matches = response.get('UserMatches', [])
        if not matches:
            return make_response(jsonify({"status": "failed", "message": "No match found"}), 200)

        user_id = matches[0]['User'].get('UserId')
        similarity = matches[0]['Similarity']

        return make_response(jsonify({
            "status": "success",
            "user_id": user_id,
            "message": f"Welcome back, {user_id}!",
            "confidence": f"{similarity:.2f}%"
        }), 200)

    except Exception as e:
        return make_response(jsonify({"error": str(e)}), 400)


# ─── Legacy Export (kept for backward compat) ─────────────────────────────────

@app_views.route('/export-attendance', methods=["POST"], strict_slashes=False)
@jwt_required()
@role_required('lecturer', 'admin')
def export_attendance():
    """Export attendance data to Excel."""
    data = request.get_json()
    if not data:
        return make_response(jsonify({"error": "No data provided"}), 400)

    session_id = data.get("session_id")
    session_obj = storage.get_id(Sessions, session_id) if session_id else None
    course = ses.query(Courses).filter_by(courseID=session_obj.courseID).first() if session_obj else None

    session_info = {
        "course_code": session_obj.courseID if session_obj else "N/A",
        "course_name": course.courseName if course else "N/A",
        "session_name": session_obj.session_name if session_obj else "N/A",
    }

    grouped = data.get("grouped_users", data)
    return export_attendance_to_excel(grouped, session_info)


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def get_user(user_dict, session_id):
    """
    Look up user details and attendance info for recognized faces.
    Groups results by department.

    Args:
        user_dict: list of {"user_id": ..., "confidence": ...}
        session_id: the session to look up attendance for (REQUIRED)
    """
    grouped_users = {}

    for user_entry in user_dict:
        user_id = user_entry.get("user_id")
        user_obj = storage.get_id(User, user_id)

        if not user_obj:
            continue

        dept = "unknown"
        if user_obj.university and user_obj.university.department:
            dept = user_obj.university.department.lower()

        att_record = ses.query(Attendance).filter(
            Attendance.user_id == user_id,
            Attendance.session_id == session_id
        ).first()

        student_info = {
            "user_id": user_id,
            "name": user_obj.name,
            "matric": user_obj.matric,
            "confidence": user_entry.get("confidence", 0),
            "days": att_record.days if att_record else 0,
            "percentage": att_record.percentage if att_record else "0.00%",
            "eligibility": att_record.eligibility if att_record else "Ineligible"
        }

        if dept not in grouped_users:
            grouped_users[dept] = []
        grouped_users[dept].append(student_info)

    return grouped_users


def update_student_record(all_results, session_id, lecturer_id=None):
    """
    Record attendance for recognized students.
    Returns dict with recognized_count, skipped_count, and skipped_users list.

    Args:
        all_results: deduplicated list of {"user_id": ..., "confidence": ...}
        session_id: REQUIRED - which session to record attendance for
        lecturer_id: who recorded the attendance
    """
    REQUIRED_PERCENTAGE = 75

    session_obj = storage.get_id(Sessions, session_id)
    if not session_obj:
        return {"recognized_count": 0, "skipped_count": 0, "skipped_users": []}

    TOTAL_LECTURES = session_obj.total_expected_classes or 13

    recognized_count = 0
    skipped_count = 0
    skipped_users = []

    for user_entry in all_results:
        user_id = user_entry.get("user_id")
        if not user_id:
            continue

        # Verify user exists
        user = storage.get_id(User, user_id)
        if not user:
            continue

        # ── PRODUCTION: Per-day deduplication (uncomment for 24hr logic) ──
        # today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        # existing_log = ses.query(AttendanceLog).filter(
        #     AttendanceLog.session_id == session_id,
        #     AttendanceLog.user_id == user_id,
        #     AttendanceLog.date >= today_start
        # ).first()
        #
        # if existing_log:
        #     skipped_count += 1
        #     skipped_users.append({
        #         "user_id": user_id,
        #         "name": user.name,
        #         "reason": "Already recorded today"
        #     })
        #     continue

        # ── DEMO MODE: Every upload creates a new attendance record ──
        log = AttendanceLog(
            session_id=session_id,
            user_id=user_id,
            recorded_by=lecturer_id,
            date=datetime.utcnow(),
            status='present'
        )
        log.save()
        recognized_count += 1

        # ── Recalculate summary from actual log count ──
        _recalculate_attendance_summary(session_id, user_id)

    return {
        "recognized_count": recognized_count,
        "skipped_count": skipped_count,
        "skipped_users": skipped_users
    }


def _recalculate_attendance_summary(session_id, user_id):
    """Recalculate the cached Attendance summary record from AttendanceLogs."""
    REQUIRED_PERCENTAGE = 75

    session_obj = storage.get_id(Sessions, session_id)
    total_classes = session_obj.total_expected_classes if session_obj else 13

    days_present = ses.query(AttendanceLog).filter_by(
        session_id=session_id, user_id=user_id, status='present'
    ).count()
    days_excused = ses.query(AttendanceLog).filter_by(
        session_id=session_id, user_id=user_id, status='excused'
    ).count()

    effective = days_present + days_excused
    percentage = (effective / total_classes) * 100 if total_classes > 0 else 0

    # Update or create summary
    attendance = ses.query(Attendance).filter_by(
        session_id=session_id, user_id=user_id
    ).first()

    if attendance:
        attendance.days = days_present
        attendance.percentage = f"{percentage:.2f}%"
        attendance.eligibility = "Eligible" if percentage >= REQUIRED_PERCENTAGE else "Ineligible"
        attendance.status = 'present' if days_present > 0 else 'absent'
        attendance.save()
    else:
        attendance = Attendance(
            user_id=user_id,
            session_id=session_id,
            days=days_present,
            lecturer_id="system",
            status='present' if days_present > 0 else 'absent',
            percentage=f"{percentage:.2f}%",
            eligibility="Eligible" if percentage >= REQUIRED_PERCENTAGE else "Ineligible"
        )
        attendance.save()

    # Check and send notifications
    _check_and_notify(user_id, session_id, percentage, total_classes, days_present)


def _check_and_notify(user_id, session_id, percentage, total_lectures, days_attended):
    """Send notification to student if their attendance drops into warning zones."""
    session_obj = storage.get_id(Sessions, session_id)
    course = ses.query(Courses).filter_by(
        courseID=session_obj.courseID
    ).first() if session_obj else None
    course_label = f"{course.courseName} ({session_obj.courseID})" if course and session_obj else "your course"

    remaining = total_lectures - days_attended

    # Warning zone: 76-80%
    if 76 <= percentage <= 80:
        title = f"⚠️ Attendance Warning — {course_label}"
        message = (
            f"Your attendance is at {percentage:.1f}%. "
            f"You have {remaining} classes left. "
            f"Missing more classes may make you ineligible for exams."
        )
        _create_notification_if_new(user_id, session_id, title, message, "warning")

    # Danger zone: below 75%
    elif percentage < 75:
        title = f"🚨 Attendance Critical — {course_label}"
        message = (
            f"Your attendance has dropped to {percentage:.1f}%, "
            f"below the 75% required for exam eligibility. "
            f"Please attend all remaining classes."
        )
        _create_notification_if_new(user_id, session_id, title, message, "danger")


def _create_notification_if_new(user_id, session_id, title, message, notif_type):
    """Create a notification only if a similar one hasn't been sent today."""
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    existing = ses.query(Notification).filter(
        Notification.user_id == user_id,
        Notification.related_session_id == session_id,
        Notification.notification_type == notif_type,
        Notification.created_at >= today_start
    ).first()

    if not existing:
        notification = Notification(
            user_id=user_id,
            title=title,
            message=message,
            notification_type=notif_type,
            related_session_id=session_id,
        )
        notification.save()


def export_attendance_to_excel(grouped_users, session_info=None):
    """Export attendance data to an Excel file with dynamic header info."""
    output = io.BytesIO()

    if not grouped_users:
        grouped_users = {"Empty": []}

    info = session_info or {}

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame().to_excel(writer, sheet_name='Attendance')

        workbook = writer.book
        worksheet = writer.sheets['Attendance']
        bold_font = Font(bold=True)

        fixed_info = [
            [f"COURSE CODE: {info.get('course_code', 'N/A')}"],
            [f"COURSE NAME: {info.get('course_name', 'N/A')}"],
            [f"SESSION: {info.get('session_name', 'N/A')}"],
            [f"DATE: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
        ]
        for i, row_data in enumerate(fixed_info, 1):
            worksheet.cell(row=i, column=1, value=row_data[0]).font = bold_font

        start_row = 6
        for dept, students in grouped_users.items():
            dept_cell = worksheet.cell(row=start_row + 1, column=1)
            dept_cell.value = dept.upper()
            dept_cell.font = bold_font

            if students:
                df = pd.DataFrame(students)
                cols = [c for c in ["name", "matric", "days", "percentage", "eligibility"] if c in df.columns]
                if cols:
                    df = df[cols]
                    df.columns = [c.upper().replace("_", " ") for c in cols]
                df.index = df.index + 1
                df.to_excel(writer, startrow=start_row + 1, sheet_name='Attendance')
                start_row += len(students) + 4
            else:
                worksheet.cell(row=start_row + 2, column=1, value="No students found")
                start_row += 4

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Attendance_Report_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )